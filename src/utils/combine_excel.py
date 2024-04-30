"""Combine multiple excel files into one.

Usage:
    python combine_excel.py <target> <file1> <file2> ...
"""
import sys
from collections.abc import Iterable
from openpyxl import load_workbook, Workbook


def merge_excel_files(filepaths: Iterable[str], target: str) -> None:
    """Merge multiple excel files into one."""
    merged_workbook = Workbook()
    merged_workbook.remove(merged_workbook.active)

    for file in filepaths:
        workbook = load_workbook(file)
        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            merged_sheet = merged_workbook.create_sheet(title=sheet_name)
            for row in sheet.rows:
                merged_sheet.append([cell.value for cell in row])
    
    merged_workbook.save(target)


if __name__ == "__main__":
    target = sys.argv[1]
    files = sys.argv[2:]
    merge_excel_files(files, target)
